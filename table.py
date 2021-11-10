from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# A contact can have a first name, last name, and a phone number
class Contact:
    def __init__(self, fname, lname, phone):
        self.fname = fname
        self.lname = lname
        self.phone = phone

# Find out how many contacts the user will be adding
num_contacts = int(input("How many contacts will you be adding? "))

# List containing Contact objects
contact_list = []

# Get the infomation for each contact from the user and add it to excel
for c in range (num_contacts):
    fname = input(f"Enter the first name of contact {c + 1}: ")
    lname = input(f"Enter the last name of contact {c + 1}: ")
    phone = input(f"Enter the phone number of contact {c + 1}: ")
    contact_list.append(Contact(fname, lname, phone))

# # Print all contacts
# for contact in contact_list:
#     print(contact.fname)
#     print(contact.lname)
#     print(contact.phone)

# Initialize a workbook
wb = Workbook()

# Create an active work sheet
ws = wb.active

# Title of the worksheet
ws.title = "Contact List"

# Table headings
ws.append(["First Name", "Last Name", "Phone number"])

# Print all contacts and their information
for contact in contact_list:
    ws.append([contact.fname, contact.lname, contact.phone])

print("Contacts successfully added to Excel!")

# for row in range(1, 11):
#     for col in range(1,5):
#         char = get_column_letter(col)
#         print(ws[char + str(row)])

# ws.insert_rows(7)
# ws.delete_rows(7)

# Save the workbook
wb.save("contacts.xlsx")

